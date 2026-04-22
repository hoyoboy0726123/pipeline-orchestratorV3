"""
Pipeline 狀態機主引擎。

流程：
  START → 逐步執行 → LLM 驗證 → 通過則下一步
                                 → 失敗且有重試次數 → 自動重試
                                 → 失敗且重試耗盡  → 暫停 + Telegram inline keyboard
  用戶按 [重試 / 跳過 / 中止] → resume_pipeline() 繼續或結束

Telegram 通知時機：
  - 步驟失敗需人為決策 → 詢問訊息 + inline keyboard
  - Pipeline 全部完成 / 中止 → 結果摘要
"""
import asyncio
import logging
import uuid
from datetime import datetime
from typing import Optional

from telegram import Bot, InlineKeyboardButton, InlineKeyboardMarkup

from config import TELEGRAM_BOT_TOKEN, TELEGRAM_CHAT_ID
from .models import PipelineConfig
from .store import PipelineRun, StepResult, get_store
from .logger import create_run_logger, resume_run_logger
from .executor import execute_step, execute_step_with_skill
from .validator import validate_step, validate_step_with_skill, ValidationResult


# ── Abort flags（in-memory）────────────────────────────────────────────────────
_abort_flags: set[str] = set()

# ── Running task tracking（for immediate cancel）──────────────────────────────
_running_tasks: dict[str, asyncio.Task] = {}


def register_task(run_id: str, task: asyncio.Task):
    _running_tasks[run_id] = task


def unregister_task(run_id: str):
    _running_tasks.pop(run_id, None)


def request_abort(run_id: str):
    """前端/API 呼叫：標記此 run 需要中止"""
    _abort_flags.add(run_id)


async def force_abort(run_id: str):
    """立即中止：kill 子進程 + 標記 computer_use abort + cancel asyncio task + 更新狀態"""
    from .executor import kill_run_processes
    from .computer_use import request_abort as _cu_abort
    _abort_flags.add(run_id)
    # 1. 立即 kill 所有子進程
    kill_run_processes(run_id)
    # 1a. 通知 computer_use 引擎中止（它跑在 executor thread 裡，kill 不到）
    _cu_abort(run_id)
    # 2. Cancel asyncio task
    task = _running_tasks.pop(run_id, None)
    if task and not task.done():
        task.cancel()
        try:
            await task
        except (asyncio.CancelledError, Exception):
            pass
    # 3. 更新 run 狀態
    store = get_store()
    run = store.load(run_id)
    if run and run.status in ("running", "awaiting_human"):
        run.status = "aborted"
        run.ended_at = datetime.now().isoformat()
        store.save(run)
        logger = logging.getLogger(f"pipeline.{run_id}")
        logger.info("⛔ Pipeline 被立即中止（force abort）")
        try:
            config = PipelineConfig.from_dict(run.config_dict)
            await _notify_final(run, config)
        except Exception:
            pass
    clear_abort(run_id)


def is_abort_requested(run_id: str) -> bool:
    return run_id in _abort_flags


def clear_abort(run_id: str):
    _abort_flags.discard(run_id)


# ── Telegram helpers ─────────────────────────────────────────────────────────

def _decision_keyboard(run_id: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("🔄 重試", callback_data=f"pipe_retry:{run_id}"),
            InlineKeyboardButton("💬 補充指示", callback_data=f"pipe_hint:{run_id}"),
        ],
        [
            InlineKeyboardButton("📋 查看 Log", callback_data=f"pipe_log:{run_id}"),
            InlineKeyboardButton("🛑 中止", callback_data=f"pipe_abort:{run_id}"),
        ],
    ])


def _confirm_keyboard(run_id: str, screenshot: bool = False) -> InlineKeyboardMarkup:
    rows = [
        [
            InlineKeyboardButton("✅ 繼續執行", callback_data=f"pipe_continue:{run_id}"),
            InlineKeyboardButton("💬 補充指示", callback_data=f"pipe_hint:{run_id}"),
        ],
        [
            InlineKeyboardButton("📋 查看 Log", callback_data=f"pipe_log:{run_id}"),
            InlineKeyboardButton("🛑 中止", callback_data=f"pipe_abort:{run_id}"),
        ],
    ]
    if screenshot:
        rows.insert(1, [
            InlineKeyboardButton("📸 截圖", callback_data=f"pipe_screenshot:{run_id}"),
        ])
    return InlineKeyboardMarkup(rows)


def _ask_user_keyboard(run_id: str, options: list) -> InlineKeyboardMarkup:
    """
    ask_user 問題的 Telegram 鍵盤。
    - 有 options → 每個選項一個按鈕（一行最多 2 個）+ 自由輸入 + 中止
    - 無 options → 只有「自由輸入」和「中止」
    """
    rows: list[list[InlineKeyboardButton]] = []
    if options:
        # callback 長度上限 64 bytes，用索引傳遞
        row: list[InlineKeyboardButton] = []
        for i, opt in enumerate(options):
            label = str(opt)
            if len(label) > 30:
                label = label[:27] + "…"
            row.append(InlineKeyboardButton(label, callback_data=f"pipe_answer:{run_id}:{i}"))
            if len(row) == 2:
                rows.append(row)
                row = []
        if row:
            rows.append(row)
    rows.append([
        InlineKeyboardButton("✍ 自由輸入", callback_data=f"pipe_answer_free:{run_id}"),
        InlineKeyboardButton("🛑 中止", callback_data=f"pipe_abort:{run_id}"),
    ])
    return InlineKeyboardMarkup(rows)


async def _send_ask_user_notification(run, question: str, options: list, context: str, step_name: str):
    """Skill agent 呼叫 ask_user 時發送 Telegram 通知。"""
    import html as _html
    total = len(run.config_dict.get("steps", [])) if run.config_dict else 0
    step_num = run.current_step + 1
    lines = [
        "❓ <b>AI 技能請求人工協助</b>",
        "",
        f"📋 {run.pipeline_name}",
        f"📍 步驟 {step_num}/{total}：<b>{_html.escape(step_name)}</b>",
        "",
        f"<b>問題</b>：{_html.escape(question)}",
    ]
    if context:
        lines.append(f"\n<b>背景</b>：{_html.escape(context)}")
    if options:
        lines.append("\n請從下方選項選擇，或點「自由輸入」回答。")
    else:
        lines.append("\n請點「自由輸入」並傳送文字回答。")
    await _tg_send(run.telegram_chat_id, "\n".join(lines),
                   _ask_user_keyboard(run.run_id, options))


def _is_valid_tg_token(token: str) -> bool:
    """檢查 Telegram Bot Token 格式是否正確（數字:字母混合）"""
    if not token or ":" not in token:
        return False
    parts = token.split(":", 1)
    return parts[0].isdigit() and len(parts[1]) > 10


def _get_tg_token() -> str:
    """取得 Telegram Bot Token（優先用 settings UI 設定，fallback 到 env）"""
    logger = logging.getLogger("pipeline")
    try:
        from settings import get_settings
        token = get_settings().get("telegram_bot_token", "")
        if token and _is_valid_tg_token(token):
            return token
        elif token:
            logger.warning(f"[Telegram] settings 中的 token 格式不正確（'{token[:15]}...'），改用 .env")
    except Exception:
        pass
    if TELEGRAM_BOT_TOKEN:
        logger.debug(f"[Telegram] 使用 .env 的 TELEGRAM_BOT_TOKEN")
    return TELEGRAM_BOT_TOKEN


def _get_tg_chat_id() -> int:
    """取得 Telegram Chat ID（優先 settings UI，fallback 到 env）"""
    logger = logging.getLogger("pipeline")
    try:
        from settings import get_settings
        cid = get_settings().get("telegram_chat_id", "")
        if cid:
            return int(cid)
    except Exception:
        pass
    # fallback 到 .env
    if TELEGRAM_CHAT_ID:
        try:
            logger.debug(f"[Telegram] 使用 .env 的 TELEGRAM_CHAT_ID")
            return int(TELEGRAM_CHAT_ID)
        except ValueError:
            logger.warning(f"[Telegram] .env TELEGRAM_CHAT_ID 格式不正確：{TELEGRAM_CHAT_ID}")
    return 0


async def _tg_send(chat_id: int, text: str, reply_markup=None):
    """發送 Telegram 訊息（錯誤靜默記錄，不拋出）"""
    logger = logging.getLogger("pipeline")
    token = _get_tg_token()
    # 如果沒傳 chat_id，嘗試從 settings 取得
    if not chat_id:
        chat_id = _get_tg_chat_id()
    if not chat_id or not token:
        logger.warning(f"[Telegram] 跳過發送：chat_id={chat_id}, token={'有' if token else '無'}")
        return
    logger.info(f"[Telegram] 發送訊息到 chat_id={chat_id}（token={token[:15]}...）")
    try:
        bot = Bot(token=token)
        await bot.send_message(
            chat_id=chat_id,
            text=text,
            parse_mode="HTML",
            reply_markup=reply_markup,
        )
        await bot.close()
        logger.info(f"[Telegram] ✅ 發送成功")
    except Exception as e:
        logger.error(f"[Telegram] ❌ 發送失敗：{e}")


def take_screenshot(pipeline_name: str, step_name: str) -> Optional[str]:
    """截取螢幕畫面，儲存到工作流資料夾，回傳檔案路徑。"""
    import time as _t
    from pathlib import Path as _P
    logger = logging.getLogger("pipeline")
    try:
        import mss as _mss
        _PROJ_ROOT = _P(__file__).parent.parent.parent.absolute()
        ss_dir = _PROJ_ROOT / "ai_output" / pipeline_name
        ss_dir.mkdir(parents=True, exist_ok=True)
        ss_name = f"screenshot_{step_name}_{_t.strftime('%Y%m%d_%H%M%S')}.png"
        ss_path = ss_dir / ss_name
        with _mss.mss() as sct:
            sct.shot(output=str(ss_path))
        if ss_path.exists():
            logger.info(f"[{step_name}] 📸 截圖已儲存：{ss_path}")
            return str(ss_path)
        logger.warning(f"[{step_name}] 截圖失敗：檔案未產生")
    except Exception as e:
        logger.warning(f"[{step_name}] 截圖失敗：{e}")
    return None


async def _notify_failure(run: PipelineRun, val: ValidationResult, step_name: str):
    """詢問用戶如何處理失敗步驟"""
    step_num = run.current_step + 1
    total = len(PipelineConfig.from_dict(run.config_dict).steps)
    text = (
        f"⚠️ <b>Pipeline 需要決策</b>\n\n"
        f"📋 {run.pipeline_name}\n"
        f"📍 步驟 {step_num}/{total}：<b>{step_name}</b>\n\n"
        f"🔴 {val.reason}\n"
    )
    if val.suggestion:
        text += f"💡 建議：{val.suggestion}\n"
    text += "\n請選擇處理方式："
    await _tg_send(run.telegram_chat_id, text, _decision_keyboard(run.run_id))


async def _notify_final(run: PipelineRun, config: PipelineConfig):
    """發送 pipeline 最終結果摘要"""
    total = len(config.steps)
    ok_count = sum(1 for r in run.step_results if r.validation_status == "ok")

    status_map = {
        "completed": ("✅", "Pipeline 完成"),
        "aborted":   ("🛑", "Pipeline 已中止"),
    }
    emoji, title = status_map.get(run.status, ("❌", "Pipeline 失敗"))

    duration = ""
    if run.ended_at and run.started_at:
        try:
            secs = int((
                datetime.fromisoformat(run.ended_at) -
                datetime.fromisoformat(run.started_at)
            ).total_seconds())
            duration = f"⏱ 耗時：{secs // 60}m {secs % 60}s\n"
        except Exception:
            pass

    # Step 摘要
    step_lines = []
    for i, step in enumerate(config.steps):
        if i < len(run.step_results):
            r = run.step_results[i]
            icon = {"ok": "✅", "warning": "⚠️", "failed": "❌"}.get(r.validation_status, "❓")
            step_lines.append(f"  {icon} {step.name}")
        else:
            step_lines.append(f"  ⬜ {step.name}（未執行）")

    text = (
        f"{emoji} <b>{title}</b>\n\n"
        f"📋 {run.pipeline_name}\n"
        f"🔢 {ok_count}/{total} 步驟成功\n"
        f"{duration}"
        f"\n<b>步驟概覽：</b>\n" + "\n".join(step_lines) +
        f"\n\n📁 <code>{run.log_path}</code>"
    )
    await _tg_send(run.telegram_chat_id, text)


# ── Deterministic validation (fast recipe mode) ──────────────────────────────

def _deterministic_validate(step, exec_result, logger) -> ValidationResult:
    """Recipe 快速模式：不叫 LLM，只做確定性檢查。"""
    from pathlib import Path as _Path

    # 1. exit code
    if exec_result.exit_code != 0:
        return ValidationResult(
            status="failed",
            reason=f"Exit code {exec_result.exit_code}",
            suggestion="Recipe 執行失敗，建議改用完整模式重跑",
        )

    # 2. 輸出檔存在 + 大小
    if step.output and step.output.path:
        p = _Path(step.output.path).expanduser()
        if not p.exists():
            return ValidationResult(
                status="failed",
                reason=f"輸出檔案 {step.output.path} 不存在",
                suggestion="Recipe 未產生預期檔案，建議改用完整模式",
            )
        size = p.stat().st_size
        if size == 0:
            return ValidationResult(
                status="failed",
                reason=f"輸出檔案 {step.output.path} 為空檔案（0 bytes）",
                suggestion="Recipe 產生了空檔案，建議改用完整模式",
            )
        # CSV: 檢查有 header
        if p.suffix.lower() == ".csv":
            try:
                with open(p, "r", encoding="utf-8") as f:
                    lines = sum(1 for _ in f)
                if lines < 2:
                    return ValidationResult(
                        status="failed",
                        reason=f"CSV 檔案只有 {lines} 行（預期至少有 header + 資料）",
                        suggestion="",
                    )
            except Exception:
                pass
        # Excel: 檢查有 sheet
        if p.suffix.lower() in (".xlsx", ".xls"):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(p, read_only=True)
                sheet_count = len(wb.sheetnames)
                wb.close()
                if sheet_count == 0:
                    return ValidationResult(
                        status="failed",
                        reason="Excel 檔案沒有任何工作表",
                        suggestion="",
                    )
            except Exception:
                pass

    logger.info(f"[{step.name}] ⚡ Recipe 快速驗證通過（確定性檢查）")
    return ValidationResult(
        status="ok",
        reason="Recipe 快速模式：exit code=0、輸出檔案存在且非空",
        suggestion="",
    )


def get_run_log_tail(run_id: str, lines: int = 30) -> str:
    """取得 pipeline 執行 log 的最後 N 行（供 Telegram 查看）"""
    store = get_store()
    run = store.load(run_id)
    if not run or not run.log_path:
        return "（找不到 log）"
    from pathlib import Path as _Path
    log_file = _Path(run.log_path)
    if not log_file.exists():
        return "（log 檔案不存在）"
    try:
        all_lines = log_file.read_text(encoding="utf-8").splitlines()
        tail = all_lines[-lines:] if len(all_lines) > lines else all_lines
        return "\n".join(tail)
    except Exception as e:
        return f"（讀取失敗：{e}）"


# ── Main pipeline engine ──────────────────────────────────────────────────────

async def run_pipeline(
    config_dict: dict,
    chat_id: int,
    run_id: Optional[str] = None,
    start_from_step: int = 0,
) -> str:
    """
    執行（或恢復）一個 pipeline。
    """
    store = get_store()

    # 建立或恢復 run
    if run_id:
        run = store.load(run_id)
        if not run:
            raise ValueError(f"找不到 pipeline run: {run_id}")
        
        # 確保 run 物件同步使用傳入的最新配置（包含 hint）
        run.config_dict = config_dict
        run.status = "running"
        run.current_step = start_from_step
        # 附加到原始 log 檔（不建新檔），前端讀到的 log_path 保持不變
        logger = resume_run_logger(run.run_id, run.log_path)
        logger.info(f"恢復執行，從步驟 {start_from_step + 1} 繼續")
    else:
        # 新建 run
        config = PipelineConfig.from_dict(config_dict)
        run_id = str(uuid.uuid4())[:12]
        logger, log_path = create_run_logger(run_id, config.name)
        run = PipelineRun(
            run_id=run_id,
            pipeline_name=config.name,
            config_dict=config_dict,
            telegram_chat_id=chat_id,
            log_path=log_path,
        )
        logger.info(f"Pipeline 開始：{config.name}，共 {len(config.steps)} 步驟")

    config = PipelineConfig.from_dict(run.config_dict)
    use_recipe = run.config_dict.get("_use_recipe", False)
    workflow_id = run.config_dict.get("_workflow_id") or run.workflow_id
    store.save(run)

    # ── Step loop ────────────────────────────────────────────
    completed_outputs: list[dict] = []  # 收集前步驟的輸出資訊

    # 恢復執行時，重建已完成步驟的輸出資訊（供後續步驟參考）
    if start_from_step > 0:
        from pathlib import Path as _Path
        for i in range(start_from_step):
            prev_step = config.steps[i] if i < len(config.steps) else None
            if prev_step and not prev_step.human_confirm and prev_step.output and prev_step.output.path:
                p = _Path(prev_step.output.path).expanduser()
                out_info = {"path": str(p), "schema": ""}
                try:
                    if p.suffix == ".csv" and p.exists():
                        with open(p, "r") as f:
                            out_info["schema"] = f.readline().strip()
                    elif p.suffix in (".xlsx", ".xls") and p.exists():
                        out_info["schema"] = "Excel 工作簿"
                    elif p.suffix in (".png", ".jpg", ".jpeg") and p.exists():
                        out_info["schema"] = "圖片檔案"
                except Exception:
                    pass
                completed_outputs.append(out_info)
        if completed_outputs:
            logger.info(f"已重建 {len(completed_outputs)} 個前步驟的輸出資訊：{[o['path'] for o in completed_outputs]}")

    no_save_recipe = run.config_dict.get("_no_save_recipe", False)

    while run.current_step < len(config.steps):
        # ── 每步開始前檢查中止旗標 ──
        if is_abort_requested(run.run_id):
            clear_abort(run.run_id)
            unregister_task(run.run_id)
            run.status = "aborted"
            run.ended_at = datetime.now().isoformat()
            store.save(run)
            logger.info("用戶透過 UI 中止 Pipeline")
            await _notify_final(run, config)
            return run.run_id

        step = config.steps[run.current_step]
        step_num = run.current_step + 1
        total = len(config.steps)
        logger.info(f"══ 步驟 {step_num}/{total}：{step.name} ══")

        # ── 人工確認節點：暫停等待確認 ──
        if step.human_confirm:
            logger.info(f"[{step.name}] ✋ 人工確認節點，暫停等待確認")

            # 收集前一步結果摘要
            prev_summary = ""
            if run.step_results:
                prev = run.step_results[-1]
                status_icon = {"ok": "✅", "failed": "❌"}.get(prev.validation_status, "⚠️")
                prev_summary = (
                    f"前一步驟：{prev.step_name}\n"
                    f"狀態：{status_icon} {prev.validation_status}\n"
                    f"原因：{prev.validation_reason or '（無）'}\n"
                )
                if prev.stdout_tail:
                    prev_summary += f"輸出摘要：{prev.stdout_tail[-300:]}\n"

            confirm_msg = step.message or "請確認上一步結果是否正確，再繼續執行"
            full_message = f"{prev_summary}\n📋 {confirm_msg}"

            run.status = "awaiting_human"
            run.awaiting_type = "human_confirm"
            run.awaiting_message = confirm_msg
            store.save(run)

            # Telegram 通知
            if step.notify_telegram:
                tg_text = (
                    f"✋ <b>Pipeline 等待確認</b>\n\n"
                    f"📋 {run.pipeline_name}\n"
                    f"📍 步驟 {step_num}/{total}：<b>{step.name}</b>\n\n"
                )
                if prev_summary:
                    tg_text += f"{prev_summary}\n"
                tg_text += f"💬 {confirm_msg}\n\n請選擇："
                await _tg_send(run.telegram_chat_id, tg_text,
                               _confirm_keyboard(run.run_id, screenshot=step.screenshot))

            # 記錄此步驟的結果（標記為等待中）
            step_result = StepResult(
                step_index=run.current_step,
                step_name=step.name,
                exit_code=0,
                stdout_tail="等待人工確認",
                stderr_tail="",
                validation_status="ok",
                validation_reason="人工確認節點 — 等待中",
                validation_suggestion="",
                retries_used=0,
            )
            if len(run.step_results) > run.current_step:
                run.step_results[run.current_step] = step_result
            else:
                run.step_results.append(step_result)
            store.save(run)
            unregister_task(run.run_id)
            return run.run_id  # 暫停，等 resume_pipeline 被呼叫

        logger.debug(f"[{step.name}] batch 全文（{len(step.batch)} 字元）：{step.batch[:500]}")

        retries_used = 0
        step_failures: list[dict] = []  # 累積此步驟的失敗歷史，傳給下次重試

        # 計算當前步驟的工作目錄 (Working Directory)
        from pathlib import Path as _Path
        # 定義專案根目錄 (backend/pipeline/runner.py 的上三層)
        _PROJ_ROOT = _Path(__file__).parent.parent.parent.absolute()

        def _resolve_path(p: str) -> _Path:
            """把 output.path 解析成絕對路徑：
            - `~/xxx` 展開到使用者家目錄
            - 絕對路徑直接用
            - 相對路徑 → 以**專案根目錄**為基準（而非 backend cwd）"""
            pp = _Path(p).expanduser()
            if not pp.is_absolute():
                pp = _PROJ_ROOT / pp
            return pp

        # 預設：專案根目錄/ai_output/{pipeline_name}/
        default_wd = str(_PROJ_ROOT / "ai_output" / config.name)
        wd = step.working_dir
        if not wd and step.output and step.output.path:
            wd = str(_resolve_path(step.output.path).parent)
        if not wd:
            wd = default_wd
        _Path(wd).mkdir(parents=True, exist_ok=True)

        # Retry loop for this step
        while True:
            if step.computer_use:
                # ── 桌面自動化節點：純 pyautogui + opencv，不走 LLM / recipe ──
                from .computer_use import execute_computer_use_step
                from .executor import ExecResult as _ExecResult
                # assets_dir 相對路徑解析：若為空，預設 ai_output/<pipeline>/<step_name>_assets
                if step.assets_dir:
                    assets_abs = str(_resolve_path(step.assets_dir))
                else:
                    assets_abs = str(_resolve_path(f"ai_output/{config.name}/{step.name}_assets"))
                # actions 是 ComputerUseAction pydantic model，轉成 dict list 傳進引擎
                actions_dicts = [a.model_dump() if hasattr(a, "model_dump") else dict(a) for a in (step.actions or [])]
                _cu_result = await asyncio.get_event_loop().run_in_executor(
                    None,
                    lambda: execute_computer_use_step(
                        actions=actions_dicts,
                        assets_dir=assets_abs,
                        logger=logger,
                        run_id=run.run_id,
                        fail_fast=step.fail_fast,
                        cv_threshold=step.cv_threshold,
                        cv_search_only_near=step.cv_search_only_near,
                        cv_search_radius=step.cv_search_radius,
                        cv_trigger_hover=step.cv_trigger_hover,
                        cv_hover_wait_ms=step.cv_hover_wait_ms,
                        cv_coord_fallback=step.cv_coord_fallback,
                    ),
                )
                # 映射回 ExecResult 讓後續驗證/重試邏輯通用
                exec_result = _ExecResult(
                    exit_code=_cu_result.exit_code,
                    stdout=_cu_result.stdout,
                    stderr=_cu_result.stderr,
                )
            elif step.skill_mode:
                # recipe key 使用「索引:名稱」避免同名步驟互相覆蓋
                recipe_step_key = f"{step_num}:{step.name}"
                # 把 output_path 解析成絕對路徑傳給 LLM，避免 LLM 搞不清楚相對於哪個 cwd
                _resolved_out = str(_resolve_path(step.output.path)) if (step.output and step.output.path) else None
                exec_result = await execute_step_with_skill(
                    task_description=step.batch,
                    timeout=step.timeout,
                    logger=logger,
                    step_name=step.name,
                    output_path=_resolved_out,
                    working_dir=wd,
                    prev_outputs=completed_outputs if completed_outputs else None,
                    pipeline_id=workflow_id or config.name,
                    use_recipe=use_recipe,
                    no_save_recipe=no_save_recipe,
                    readonly=step.readonly,
                    run_id=run.run_id,
                    previous_failures=step_failures if step_failures else None,
                    recipe_step_key=recipe_step_key,
                    skill_name=step.skill,
                    ask_mode=step.ask_mode,
                )
            else:
                exec_result = await execute_step(
                    command=step.batch,
                    timeout=step.timeout,
                    logger=logger,
                    step_name=step.name,
                    run_id=run.run_id,
                    working_dir=wd,
                )

            # 快速模式：Recipe 命中 + 執行成功 → 確定性驗證（不叫 LLM）
            recipe_hit = (exec_result.stderr == "__RECIPE_HIT__")
            if recipe_hit:
                exec_result.stderr = ""  # 清掉標記

            has_expect = step.output and step.output.get_expect()
            # computer_use 節點：成敗已由 action 執行結果決定，不需 LLM 驗證
            if step.computer_use:
                _status = "ok" if exec_result.exit_code == 0 else "failed"
                val = ValidationResult(
                    status=_status,
                    reason=f"桌面自動化 {exec_result.stdout.count('OK')} 個動作成功"
                           + (f"，{exec_result.exit_code} 個失敗" if exec_result.exit_code != 0 else ""),
                    suggestion=exec_result.stderr if _status == "failed" else "",
                )
            elif recipe_hit and use_recipe and exec_result.exit_code == 0 and not has_expect:
                # 確定性檢查：exit code=0、輸出檔存在、檔案大小合理（無 AI 驗證節點）
                val = _deterministic_validate(step, exec_result, logger)
            elif recipe_hit and use_recipe and exec_result.exit_code == 0 and has_expect:
                # Recipe 命中但有 AI 驗證節點 → 快速 LLM 驗證（不走 Skill 深度驗證）
                logger.info(f"[{step.name}] 🔍 Recipe 命中 + 有 AI 驗證需求，走快速 LLM 驗證")
                val = await validate_step(
                    step_name=step.name,
                    command=step.batch,
                    exit_code=exec_result.exit_code,
                    stdout=exec_result.stdout,
                    stderr=exec_result.stderr,
                    output_path=(str(_resolve_path(step.output.path)) if (step.output and step.output.path) else None),
                    output_expect=step.output.get_expect() if step.output else None,
                    logger=logger,
                )
            elif config.validate:
                # 完整 LLM 驗證
                use_skill = step.output and step.output.skill_mode
                validate_fn = validate_step_with_skill if use_skill else validate_step
                val = await validate_fn(
                    step_name=step.name,
                    command=step.batch,
                    exit_code=exec_result.exit_code,
                    stdout=exec_result.stdout,
                    stderr=exec_result.stderr,
                    output_path=(str(_resolve_path(step.output.path)) if (step.output and step.output.path) else None),
                    output_expect=step.output.get_expect() if step.output else None,
                    logger=logger,
                )
            else:
                status = "ok" if exec_result.exit_code == 0 else "failed"
                val = ValidationResult(
                    status=status,
                    reason=f"Exit code {exec_result.exit_code}（LLM 驗證已停用）",
                    suggestion="" if status == "ok" else "請查看 log 取得詳細錯誤",
                )
                logger.info(f"[{step.name}] 驗證（僅 exit code）：{val.status}")

            step_result = StepResult(
                step_index=run.current_step,
                step_name=step.name,
                exit_code=exec_result.exit_code,
                stdout_tail=exec_result.stdout[-500:],
                stderr_tail=exec_result.stderr[-200:],
                validation_status=val.status,
                validation_reason=val.reason,
                validation_suggestion=val.suggestion,
                retries_used=retries_used,
            )

            # 更新或追加步驟結果
            if len(run.step_results) > run.current_step:
                run.step_results[run.current_step] = step_result
            else:
                run.step_results.append(step_result)
            store.save(run)

            if val.status == "ok":
                logger.info(f"步驟 {step_num} ✅ 通過")
                # 收集延遲儲存的 recipe
                if hasattr(exec_result, 'pending_recipe') and exec_result.pending_recipe:
                    run.pending_recipes.append(exec_result.pending_recipe)
                # 收集此步驟的輸出資訊供後續步驟參考
                if step.output and step.output.path:
                    out_info = {"path": step.output.path, "schema": ""}
                    try:
                        from pathlib import Path as _Path
                        p = _Path(step.output.path)
                        if p.suffix == ".csv" and p.exists():
                            with open(p, "r") as f:
                                header = f.readline().strip()
                            out_info["schema"] = header
                        elif p.suffix in (".xlsx", ".xls") and p.exists():
                            out_info["schema"] = "Excel 工作簿"
                        elif p.suffix in (".png", ".jpg", ".jpeg") and p.exists():
                            out_info["schema"] = "圖片檔案"
                    except Exception:
                        pass
                    completed_outputs.append(out_info)
                run.current_step += 1
                store.save(run)
                break  # 進入下一步

            elif retries_used < step.retry:
                retries_used += 1
                # 記錄此次失敗的原因與建議，供下次重試時傳給 LLM
                step_failures.append({
                    "attempt": retries_used,
                    "reason": val.reason,
                    "suggestion": val.suggestion,
                    "stdout_tail": exec_result.stdout[-800:] if exec_result.stdout else "",
                    "stderr_tail": exec_result.stderr[-400:] if exec_result.stderr else "",
                })
                logger.warning(
                    f"步驟 {step_num} 驗證失敗，自動重試 {retries_used}/{step.retry}：{val.reason}"
                )
                continue  # 重試

            else:
                # 重試耗盡，暫停等待人為決策
                logger.warning(f"步驟 {step_num} 失敗且重試次數耗盡，等待人為決策")
                run.status = "awaiting_human"
                run.awaiting_type = "failure"
                run.awaiting_message = val.reason or ""

                # 優先使用 LLM 回報的 missing_packages 建立具體安裝建議
                missing_pkgs = getattr(exec_result, 'missing_packages', None) or []
                # 也嘗試從 stderr 偵測 ModuleNotFoundError
                if not missing_pkgs and exec_result.stderr:
                    import re as _re
                    found = _re.findall(r"ModuleNotFoundError: No module named '([^']+)'", exec_result.stderr)
                    if found:
                        missing_pkgs = list(dict.fromkeys(found))  # 去重保序

                if missing_pkgs:
                    pkgs_str = "、".join(missing_pkgs)
                    install_hint = f"建議在「設定 → 套件管理」安裝以下套件後再重試：{pkgs_str}"
                    run.awaiting_suggestion = install_hint + (f"\n\nAI 說明：{val.suggestion}" if val.suggestion else "")
                else:
                    run.awaiting_suggestion = val.suggestion or ""

                store.save(run)
                await _notify_failure(run, val, step.name)
                unregister_task(run.run_id)
                return run.run_id  # 暫停

    # ── 全部步驟完成 ─────────────────────────────────────────
    clear_abort(run.run_id)
    unregister_task(run.run_id)
    run.status = "completed"
    run.ended_at = datetime.now().isoformat()
    store.save(run)
    logger.info(f"Pipeline {config.name} 全部完成！")
    await _notify_final(run, config)
    return run.run_id


# ── Human-in-the-loop resume ─────────────────────────────────────────────────

async def resume_pipeline(run_id: str, decision: str, hint: str = "") -> str:
    """
    用戶透過 Telegram inline keyboard 做出決策後，呼叫此函式繼續執行。

    Args:
        run_id:   pipeline run id
        decision: "retry" | "skip" | "abort" | "continue" | "retry_with_hint"
        hint:     補充指示（retry_with_hint 時使用）

    Returns:
        str 回應訊息（回覆給用戶）
    """
    store = get_store()
    run = store.load(run_id)

    if not run:
        return f"❌ 找不到 Pipeline run：{run_id}"
    if run.status != "awaiting_human":
        return f"⚠️ Pipeline {run_id} 目前狀態為 {run.status}，無需決策"

    config = PipelineConfig.from_dict(run.config_dict)
    step_num = run.current_step + 1
    total = len(config.steps)
    # 附加到原始 log 檔，確保前端讀到的 log_path 始終指向同一個檔案
    logger = resume_run_logger(run.run_id, run.log_path)

    # ── ask_user 回答：skill agent 仍在 in-memory 等待 event ──
    if run.awaiting_type == "ask_user":
        from pipeline.executor import deliver_ask_user_answer
        if decision == "answer":
            ok = deliver_ask_user_answer(run_id, hint)
            if not ok:
                # agent 可能已 timeout 或後端已重啟
                return "⚠️ 答案送達失敗：skill agent 已不在等待狀態（可能逾時或後端重啟）"
            logger.info(f"[ask_user] 使用者答案已送達：{hint[:100]}")
            return f"✅ 答案已送出"
        elif decision == "abort":
            # 先中止 skill agent 的等待（讓它收到 None），再把 pipeline 標為 aborted
            deliver_ask_user_answer(run_id, "")  # 空字串讓 agent 繼續但不拿到答案
            run.status = "aborted"
            run.ended_at = datetime.now().isoformat()
            store.save(run)
            logger.info("[ask_user] 使用者選擇中止")
            await _notify_final(run, config)
            return f"🛑 Pipeline 已中止"
        else:
            return f"⚠️ ask_user 只接受 answer 或 abort，收到 {decision}"

    if decision == "abort":
        run.status = "aborted"
        run.ended_at = datetime.now().isoformat()
        store.save(run)
        logger.info("用戶選擇中止 Pipeline")
        await _notify_final(run, config)
        return f"🛑 Pipeline 已中止（步驟 {step_num}/{total}）"

    elif decision == "skip":
        logger.info(f"用戶選擇跳過步驟 {step_num}")
        next_step = run.current_step + 1

        if next_step >= total:
            run.status = "completed"
            run.ended_at = datetime.now().isoformat()
            store.save(run)
            await _notify_final(run, config)
            return f"⏩ 跳過最後一步，Pipeline 完成"

        run.awaiting_type = ""
        run.awaiting_message = ""
        run.awaiting_suggestion = ""
        run.status = "running"
        store.save(run)

        async def _delayed_skip():
            await asyncio.sleep(0.2)
            t = asyncio.create_task(run_pipeline(
                config_dict=run.config_dict,
                chat_id=run.telegram_chat_id,
                run_id=run.run_id,
                start_from_step=next_step,
            ))
            register_task(run.run_id, t)

        asyncio.create_task(_delayed_skip())
        return f"⏩ 跳過步驟 {step_num}，繼續執行步驟 {step_num + 1}/{total}"

    elif decision == "retry":
        logger.info(f"用戶選擇重試步驟 {step_num}")
        run.awaiting_type = ""
        run.awaiting_message = ""
        run.awaiting_suggestion = ""
        run.status = "running"
        store.save(run)

        async def _delayed_retry():
            await asyncio.sleep(0.2)
            t = asyncio.create_task(run_pipeline(
                config_dict=run.config_dict,
                chat_id=run.telegram_chat_id,
                run_id=run.run_id,
                start_from_step=run.current_step,
            ))
            register_task(run.run_id, t)

        asyncio.create_task(_delayed_retry())
        return f"🔄 重試步驟 {step_num}/{total}"

    elif decision == "retry_with_hint":
        import copy
        # 1. 使用深拷貝，確保 config 修改是獨立且完整的
        config_d = copy.deepcopy(run.config_dict)
        steps = config_d.get("steps", [])

        is_confirm = run.awaiting_type == "human_confirm"
        target = run.current_step

        if is_confirm:
            prev_step = run.current_step - 1
            while prev_step >= 0 and steps[prev_step].get("human_confirm"):
                prev_step -= 1
            if prev_step < 0:
                return "⚠️ 確認節點前沒有可重做的步驟"
            target = prev_step

        if target < len(steps):
            original_batch = steps[target].get("batch", "")
            # 清理舊的提示詞標籤，避免重複疊加
            clean_batch = original_batch.split("【用戶補充指示】")[0].strip()
            steps[target]["batch"] = f"{clean_batch}\n\n【用戶補充指示】{hint}"
            config_d["steps"] = steps

        # 2. 更新 run 狀態並「立即」同步回資料庫
        run.config_dict = config_d
        run.awaiting_type = ""
        run.awaiting_message = ""
        run.awaiting_suggestion = ""
        run.status = "running"
        store.save(run)

        # 3. 關鍵修正：給 Windows 一點點時間釋放資料庫鎖定
        async def _delayed_start():
            await asyncio.sleep(0.2)
            t = asyncio.create_task(run_pipeline(
                config_dict=config_d,  # 傳入已修改的配置
                chat_id=run.telegram_chat_id,
                run_id=run.run_id,
                start_from_step=target,
            ))
            register_task(run.run_id, t)

        asyncio.create_task(_delayed_start())
        
        if is_confirm:
            return f"💬 已附加指示，重做步驟 {target + 1}/{total}"
        else:
            return f"💬 已附加指示，重試步驟 {step_num}/{total}"

    elif decision == "continue":
        # 人工確認通過 → 繼續下一步
        logger.info(f"用戶確認通過步驟 {step_num}，繼續執行")

        # 更新確認步驟結果
        if run.current_step < len(run.step_results):
            run.step_results[run.current_step].validation_reason = "人工確認 — 已通過"
            run.step_results[run.current_step].stdout_tail = "已確認通過"

        run.awaiting_type = ""
        run.awaiting_message = ""
        next_step = run.current_step + 1

        if next_step >= total:
            run.status = "completed"
            run.ended_at = datetime.now().isoformat()
            store.save(run)
            logger.info(f"Pipeline {run.pipeline_name} 全部完成！")
            await _notify_final(run, config)
            return f"✅ 確認通過，Pipeline 全部完成"

        run.status = "running"
        store.save(run)

        async def _delayed_continue():
            await asyncio.sleep(0.2)
            t = asyncio.create_task(run_pipeline(
                config_dict=run.config_dict,
                chat_id=run.telegram_chat_id,
                run_id=run.run_id,
                start_from_step=next_step,
            ))
            register_task(run.run_id, t)

        asyncio.create_task(_delayed_continue())
        return f"✅ 確認通過，繼續執行步驟 {next_step + 1}/{total}"

    return "❓ 未知決策"
